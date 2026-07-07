#!/usr/bin/env python3
"""
update_dashboard.py - 从Excel底稿提取数据，生成JSON供gen_dashboard.py使用

用法: python update_dashboard.py "C:/path/to/7月杯数达成看板-墨柠.xlsx"
"""

import sys
import os
import json
import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# ===== 门店架构手动覆盖 =====
# 当门店架构表中缺失或需要修正某门店的dm/sm时，在此字典中添加覆盖项
# 在Excel门店架构表补齐后可移除对应条目
STORE_OVERRIDES = {
    "G73100207": {"dm": "李曦", "sm": "何海云", "store_name": "古德墨柠湖南省博物馆四楼店"},
}


def parse_date(val):
    """解析日期值，支持datetime对象和字符串"""
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        # 尝试常见格式
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.datetime.strptime(val.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def parse_date_obj(val):
    """解析日期值为datetime.date对象"""
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def safe_int(val):
    """安全转换为整数"""
    if val is None:
        return 0
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return 0


def safe_float(val):
    """安全转换为浮点数"""
    if val is None:
        return 0.0
    try:
        return float(str(val))
    except (ValueError, TypeError):
        return 0.0


def compute_yoy_date(current_date):
    """计算同比日期：去年同日，如果星期不一致则调整到最近的相同星期那天

    例：2026/7/1是周三，2025/7/1是周二（星期不一致）→ 取2025/7/2（周三）
    例：2026/6/29是周一，2025/6/29是周日 → 取2025/6/30（周一，最近方向）
    """
    try:
        same_date_last_year = current_date.replace(year=current_date.year - 1)
    except ValueError:
        # 闰年2/29 → 去年取2/28
        same_date_last_year = current_date.replace(year=current_date.year - 1, day=28)

    if same_date_last_year.weekday() == current_date.weekday():
        return same_date_last_year

    # 星期不一致，取最近方向（最多 ±3 天）调整到相同星期
    diff = current_date.weekday() - same_date_last_year.weekday()
    if diff > 3:
        diff -= 7
    elif diff < -3:
        diff += 7
    yoy_date = same_date_last_year + datetime.timedelta(days=diff)
    return yoy_date


def compute_wow_date(current_date):
    """计算环比日期：上周同期（当前日期减7天）"""
    return current_date - datetime.timedelta(days=7)


WEEKDAY_CN = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]


def compute_daily_mapping(available_dates_current):
    """根据当前月数据日期，自动计算同比和环比对比日期"""
    daily_map = {}

    for date_str in available_dates_current:
        date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        yoy_date = compute_yoy_date(date_obj)
        wow_date = compute_wow_date(date_obj)

        daily_map[date_str] = {
            "weekday": WEEKDAY_CN[date_obj.weekday()],
            "yoy_date": yoy_date.strftime("%Y-%m-%d"),
            "wow_date": wow_date.strftime("%Y-%m-%d"),
        }

    return daily_map


def read_weekly_mapping(wb):
    """读取同比环比参考日期Sheet中的周度范围数据（可选，未使用）"""
    if "同比环比参考日期" not in wb.sheetnames:
        return []

    ws = wb["同比环比参考日期"]
    weekly_map = []

    # 列F-H: 本期, 同比, 环比 (周度范围)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        period_str = str(row[5]) if row[5] else None
        if period_str and "-" in period_str:
            yoy_period_str = str(row[6]) if row[6] else None
            wow_period_str = str(row[7]) if row[7] else None
            weekly_map.append({
                "period": period_str,
                "yoy_period": yoy_period_str,
                "wow_period": wow_period_str,
            })

    return weekly_map


def compute_weekly_mapping(available_dates_current, start_date, end_date):
    """根据当前月数据日期，自动生成统计周、同比周、环比周范围

    同比周：去年同周（起始日和结束日分别用 compute_yoy_date 按星期对齐，与日度逻辑一致）
    环比周：上周（起始日-7天，结束日-7天）

    返回 [{period, yoy_period, wow_period}, ...]
    """
    if not available_dates_current:
        return []

    # 找到第一个周一
    first = start_date
    while first.weekday() != 0:  # 0=Monday
        first -= datetime.timedelta(days=1)

    weekly_map = []
    current_start = first
    while current_start <= end_date:
        current_end = current_start + datetime.timedelta(days=6)
        # 同比周：起始日和结束日分别按星期对齐（与日度同比逻辑一致）
        yoy_start = compute_yoy_date(current_start)
        yoy_end = compute_yoy_date(current_end)
        # 环比周：上周同期
        wow_start = current_start - datetime.timedelta(days=7)
        wow_end = current_end - datetime.timedelta(days=7)

        def fmt(d):
            return f"{d.year}/{d.month}/{d.day}"

        weekly_map.append({
            "period": f"{fmt(current_start)}-{fmt(current_end)}",
            "yoy_period": f"{fmt(yoy_start)}-{fmt(yoy_end)}",
            "wow_period": f"{fmt(wow_start)}-{fmt(wow_end)}",
        })
        current_start += datetime.timedelta(days=7)

    return weekly_map


def parse_weekly_range(range_str):
    """解析周度日期范围字符串如 '2026/7/5-2026/7/11'"""
    if not range_str or "-" not in range_str:
        return None, None
    parts = range_str.split("-")
    start_str = parts[0].strip()
    end_str = parts[1].strip()

    # 处理跨年的情况如 '2025/6/30-2026/7/6'
    # 格式可能是 YYYY/M/D 或 YYYY/MM/DD
    def parse_slash_date(s):
        s = s.strip()
        segments = s.split("/")
        if len(segments) == 3:
            y, m, d = int(segments[0]), int(segments[1]), int(segments[2])
            return datetime.date(y, m, d)
        return None

    start = parse_slash_date(start_str)
    end = parse_slash_date(end_str)
    return start, end


def read_store_structure(wb):
    """读取门店架构表，返回 {store_code: {dm, sm, store_name}}"""
    ws = wb["门店架构表"]
    # 列结构: 门店编码, 门店名称, 区域(区经理), 大店(大店长)
    store_map = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4, values_only=True):
        store_code = str(row[0]) if row[0] else ""
        store_name = str(row[1]) if row[1] else ""
        dm = str(row[2]) if row[2] else ""
        sm = str(row[3]) if row[3] else ""
        if not store_code:
            continue
        store_map[store_code] = {
            "store_name": store_name,
            "dm": dm,
            "sm": sm,
        }
    return store_map


def read_daily_sheet(ws):
    """读取日度数据Sheet (2025年7月底稿 / 2026年6月底稿 / 2026年7月底稿)
    列结构: 日期, 门店编码, 门店名称, 堂食饮品杯数, 外卖饮品杯数, 合计饮品杯数
    """
    data = defaultdict(dict)  # {date_str: {store_code: {dine_in, delivery, total, store_name}}}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=6, values_only=True):
        date_str = parse_date(row[0])
        store_code = str(row[1]) if row[1] else ""
        store_name = str(row[2]) if row[2] else ""
        dine_in = safe_int(row[3])
        delivery = safe_int(row[4])
        total = safe_int(row[5])

        if not date_str or not store_code:
            continue

        data[date_str][store_code] = {
            "store_name": store_name,
            "dine_in": dine_in,
            "delivery": delivery,
            "total": total,
        }

    return data


def read_daily_target(wb):
    """读取「7月分天杯数目标」Sheet

    表结构: 日期|门店编码|门店名称|杯数目标
    返回: {日期字符串(YYYY/M/D): {门店编码: 目标杯数}}
    """
    target_sheet_name = None
    for name in wb.sheetnames:
        if "分天" in name and "目标" in name:
            target_sheet_name = name
            break
    if not target_sheet_name:
        return {}

    ws = wb[target_sheet_name]
    data = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4, values_only=True):
        dt = row[0]
        store_code = str(row[1]) if row[1] else ""
        target = safe_int(row[3])
        if not dt or not store_code:
            continue
        # 标准化日期为 YYYY-MM-DD 格式（与 current_data / date_mapping key 对齐）
        if hasattr(dt, "strftime") and hasattr(dt, "year"):
            date_key = f"{dt.year:04d}-{dt.month:02d}-{dt.day:02d}"
        else:
            date_key = str(dt)
        if date_key not in data:
            data[date_key] = {}
        data[date_key][store_code] = target
    return data


def read_monthly_summary(wb):
    """读取杯数达成底稿"""
    ws = wb["杯数达成底稿"]
    # Row 2 is header: 区经理, 大店长, 门店编码, 门店, 营业天数(5月), 堂食杯数(5月), 外卖杯数(5月), 总杯数(5月),
    #                    营业天数(6月), 堂食杯数(6月), 外卖杯数(6月), 总杯数(6月),
    #                    营业天数(7月目标), 总杯数(7月目标), 总杯数(7月实际), 同比差异, 环比差异
    stores = []

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=17, values_only=True):
        dm = str(row[0]) if row[0] else ""
        sm = str(row[1]) if row[1] else ""
        store_code = str(row[2]) if row[2] else ""
        store_name = str(row[3]) if row[3] else ""

        if not store_code:
            continue

        stores.append({
            "dm": dm,
            "sm": sm,
            "store_code": store_code,
            "store_name": store_name,
            "may_days": safe_int(row[4]),
            "may_dine_in": safe_int(row[5]),
            "may_delivery": safe_int(row[6]),
            "may_total": safe_int(row[7]),
            "jun_days": safe_int(row[8]),
            "jun_dine_in": safe_int(row[9]),
            "jun_delivery": safe_int(row[10]),
            "jun_total": safe_int(row[11]),
            "jul_target_days": safe_int(row[12]),
            "jul_target_total": safe_int(row[13]),
            "jul_actual_total": safe_int(row[14]),
            "yoy_diff": safe_float(row[15]),
            "wow_diff": safe_float(row[16]),
        })

    return stores


def extract_data(xlsx_path):
    """从Excel提取所有数据，生成JSON"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 1. 读取日度数据
    sheet_names = wb.sheetnames

    # 2025年7月底稿
    yoy_data = {}
    for name in sheet_names:
        if "2025" in name and "7" in name and "底稿" in name:
            yoy_data = read_daily_sheet(wb[name])
            break

    # 2026年6月底稿
    wow_data = {}
    for name in sheet_names:
        if "2026" in name and "6" in name and "底稿" in name:
            wow_data = read_daily_sheet(wb[name])
            break

    # 2026年7月底稿 (当前月 - 可能不存在)
    current_data = {}
    current_sheet_name = None
    demo_mode = False
    for name in sheet_names:
        # 精确匹配：包含"2026"和"7月"（避免匹配到"6月"中的"7"）
        if name.startswith("2026") and ("7月" in name or "7月底稿" in name):
            current_data = read_daily_sheet(wb[name])
            current_sheet_name = name
            break

    # 如果没有7月底稿，使用6月底稿数据做演示（日期映射）
    if not current_data and wow_data:
        demo_mode = True
        print("[!] 未找到2026年7月底稿Sheet，使用6月数据做演示")
        # 将6月数据按日期映射到7月：6/1→7/1, 6/2→7/2, ... 6/30→7/30
        for jun_date_str, stores in wow_data.items():
            # 解析6月日期
            jun_date = parse_date_obj(jun_date_str)
            if jun_date is None:
                continue
            # 映射到7月同日号
            jul_date = jun_date.replace(month=7, year=2026)
            jul_date_str = jul_date.strftime("%Y-%m-%d")
            current_data[jul_date_str] = stores

    # 3. 自动计算日度对比日期（环比=上周同期，同比=去年同周期调整星期）
    available_dates_current = sorted(current_data.keys())
    daily_map = compute_daily_mapping(available_dates_current)

    # 3b. 自动生成周度映射（同比=去年同周，环比=上周同期）
    weekly_start = datetime.date(2026, 6, 29)
    weekly_end = datetime.date(2026, 9, 30)
    weekly_map = compute_weekly_mapping(available_dates_current, weekly_start, weekly_end)

    # 4. 读取门店架构表（区经理/大店长映射的权威来源）
    store_structure = read_store_structure(wb)

    # 5. 构建门店基本信息字典（优先从门店架构表获取）
    store_info = {}
    for code, info in store_structure.items():
        store_info[code] = {
            "dm": info["dm"],
            "sm": info["sm"],
            "store_name": info["store_name"],
        }

    # 门店架构表中没有的门店，从日度数据补充
    for date_data in [yoy_data, wow_data, current_data]:
        for date_str, stores in date_data.items():
            for code, info in stores.items():
                if code not in store_info:
                    store_info[code] = {
                        "dm": "",
                        "sm": "",
                        "store_name": info.get("store_name", ""),
                    }

    # 5b. 应用门店架构手动覆盖（最高优先级）
    for code, override in STORE_OVERRIDES.items():
        if code in store_info:
            for key in ["dm", "sm", "store_name"]:
                if override.get(key):
                    store_info[code][key] = override[key]
        else:
            store_info[code] = {
                "dm": override.get("dm", ""),
                "sm": override.get("sm", ""),
                "store_name": override.get("store_name", ""),
            }

    # 6. 构建区经理和大店长列表
    dm_list = sorted(set(s["dm"] for s in store_info.values() if s["dm"]))
    sm_list = sorted(set(s["sm"] for s in store_info.values() if s["sm"]))
    sm_by_dm = defaultdict(list)
    for code, info in store_info.items():
        if info["dm"] and info["sm"]:
            sm_by_dm[info["dm"]].append(info["sm"])
    # 去重排序
    for dm in sm_by_dm:
        sm_by_dm[dm] = sorted(set(sm_by_dm[dm]))

    # 7. 确定可用日期列表 (从日度数据中找出有数据的日期)
    available_dates_current = sorted(current_data.keys())
    available_dates_yoy = sorted(yoy_data.keys())
    available_dates_wow = sorted(wow_data.keys())

    # 8. 生成JSON输出
    output = {
        "meta": {
            "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "xlsx_path": xlsx_path,
            "current_sheet_found": current_sheet_name is not None,
            "demo_mode": demo_mode,
            "demo_note": "当前使用6月数据映射为7月演示数据，添加「2026年7月底稿」Sheet后重新运行脚本即可使用真实数据" if demo_mode else None,
        },
        "date_mapping": daily_map,
        "weekly_mapping": weekly_map,
        "current_data": {k: dict(v) for k, v in current_data.items()},
        "yoy_data": {k: dict(v) for k, v in yoy_data.items()},
        "wow_data": {k: dict(v) for k, v in wow_data.items()},
        "store_info": store_info,
        "dm_list": dm_list,
        "sm_list": sm_list,
        "sm_by_dm": dict(sm_by_dm),
        "monthly_summary": read_monthly_summary(wb),
        "available_dates_current": available_dates_current,
        "available_dates_yoy": available_dates_yoy,
        "available_dates_wow": available_dates_wow,
        "daily_target": read_daily_target(wb),
    }

    return output


def main():
    if len(sys.argv) < 2:
        print("用法: python update_dashboard.py <xlsx_path>")
        print("示例: python update_dashboard.py \"C:/Users/47/Desktop/7月杯数达成看板-墨柠.xlsx\"")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        print(f"文件不存在: {xlsx_path}")
        sys.exit(1)

    print(f"正在读取: {xlsx_path}")
    data = extract_data(xlsx_path)

    # 输出JSON文件 (与脚本同目录)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(script_dir, "dashboard_data.json")

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"数据提取完成，JSON已保存至: {json_path}")
    print(f"  - 日期映射: {len(data['date_mapping'])} 天")
    print(f"  - 当前月数据日期: {len(data['available_dates_current'])} 天")
    print(f"  - 去年同期数据日期: {len(data['available_dates_yoy'])} 天")
    print(f"  - 上月环比数据日期: {len(data['available_dates_wow'])} 天")
    print(f"  - 门店数: {len(data['store_info'])}")
    print(f"  - 区经理: {len(data['dm_list'])} 人")

    # 自动调用gen_dashboard.py生成HTML（直接导入调用，避免subprocess编码问题）
    gen_script = os.path.join(script_dir, "gen_dashboard.py")
    if os.path.exists(gen_script):
        print(f"\n正在生成看板HTML...")
        try:
            # 直接导入gen_dashboard模块
            import importlib.util
            spec = importlib.util.spec_from_file_location("gen_dashboard", gen_script)
            gen_mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(gen_mod)
            html_path = gen_mod.generate_html(data, os.path.join(script_dir, "7月杯数看板.html"))
            print(f"看板已生成: {html_path}")
        except Exception as e:
            print(f"生成看板失败: {e}")
            # 回退方案：直接运行脚本
            import subprocess
            try:
                result = subprocess.run(
                    [sys.executable, gen_script, json_path],
                    capture_output=True
                )
                if result.returncode == 0:
                    print(f"看板已生成")
                else:
                    print(f"生成看板失败")
            except Exception as e2:
                print(f"生成看板失败: {e2}")
    else:
        print(f"\n未找到gen_dashboard.py，请手动运行生成看板")


if __name__ == "__main__":
    main()
